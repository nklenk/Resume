# -*- coding: utf-8 -*-
import re
import openpyxl
import pandas as pd

#initialize out_files
JIL_out = open('C:\\Users\\neilklenk\\Documents\\Autosys\\Out\\JIL_out.txt', 'w')
Backout_JIL_out = open('C:\\Users\\neilklenk\\Documents\\Autosys\\Out\\Backout_JIL_out.txt', 'w')

#open the Template
with open('C:\\Users\\neilklenk\\Documents\\Autosys\\aOH JIL Template 161214.txt', 'r') as Template_handle:
    Template_contents = Template_handle.readlines()
    
#Create a dictionary of the notification email addresses
email_handle = openpyxl.load_workbook('C:\\Users\\XZ6RZM\\Documents\\Autosys\\notification emails.xlsx')
Sheet = email_handle.get_sheet_by_name('Application Email address')
email_dict = {}
for i in range(2,Sheet.max_row+1):
    email_dict[Sheet.cell(row = i, column = 1).value] = [Sheet.cell(row = i, column = 2).value]

#Initialize the lists that will be needed
acronym_job_name = []
region_ASMS_acronym_job_name = []
days_of_week = []
start_mins = []
run_window = []
email_list = []
sleep_freq = []
run_cal = []
exc_cal = []
    
#Prompt user for sheet workbook name
print("please input the name of the workbook without the extension.")
workbook = input(">")

df = pd.read_csv("C:\\Users\\neilklenk\\Documents\\Autosys\\" + workbook + ".csv", header=0)
initial_df = df

#Function to determine the index of relevant columns
def regex_counter(regex_exp, relevant_list):
    counter = 0
    for element in relevant_list:
        match = re.search(regex_exp, element)
        if match:
            break
        counter += 1
    return (counter+1)

#list building function
def list_builder(r, c):
    cell_value = str(my_sheet.cell(row = r, column = c).value)     
    indicator = 0
    my_build_list = [] 
    if cell_value == 'None':
        my_list_entry = " "
    else:
        my_str = str(cell_value)
        my_split = list(my_str)
        for char in my_split:
            if char == " ":
                indicator += 1 
        if indicator > 0:
            my_split_list = my_str.split()
            for entry in my_split_list:
                my_build_list.append(entry)
            my_full_entry = ''.join(my_build_list)
            my_list_entry = my_full_entry
        else:
            my_list_entry = cell_value
        del my_build_list[:]
    return(my_list_entry)
            
#Function to determine the sleep frequencies
def sleep_freq_calc(sleep_time_list):
    sleep_time = len(my_mins)
    sleep_freq = int(sleep_time)*60*4.75
    return(sleep_freq)

#function to create write to the out file    
def write(match_string, sub_string, relevant_list_entry, line):
    match = re.search(match_string, line)
    if match:
        if relevant_list_entry != " ":
            line = re.sub(sub_string, relevant_list_entry, line)
        else:
            line = 'skip me'
    return(line)

#get the header of the file
header = list(df)

#Dynamically assign the location of each neceeary column within the dataframe
job_name_index = regex_counter(r'\bjob_name\b', header)
app_name_index = regex_counter(r'\bapplication\b', header) #for use with the email dict
days_of_week_index = regex_counter(r'\bdays_of_week\b', header)
start_mins_index = regex_counter(r'\bstart_mins\b', header)
run_window_index = regex_counter(r'\brun_window\b', header)
run_calendar_index = regex_counter(r'\brun_calendar\b', header)
exclude_calendar_index = regex_counter(r'\bexclude_calendar\b', header)
notification_email_index = regex_counter(r'\bnotification_emailaddress\b', header)
job_type_index = regex_counter(r'\bjob_type\b', header)
date_condition_index = regex_counter(r'\bdate_conditions\b', header)

#creating a new column that contains the number of times each job starts every hour
df['start_mins_length'] = df.start_mins.str.split(',')
#Creating a boolean column saying if the job name contains "aOH" 
df['contains_aOH'] = False
for i in range(0,len(df.start_mins)):
    df.start_mins_length[i] = len(df.start_mins_length[i])
    match = re.search(r'\baOH\b', df.job_name[i])
    if match:
        df.contains_aOH[i] = True

#applying necessaty restrictions to the dataframe    
need_mon = df[df.start_mins_length > 4] 
need_mon = need_mon[need_mon.contains_aOH == False]   
need_mon = need_mon[need_mon.date_conditions != 0]
need_mon = need_mon[need_mon.job_type != 'FT']
need_mon = need_mon[need_mon.job_type != 'FW']
print('Done')

need_mon.to_excel('C:\\Users\\neilklenk\\Documents\\Autosys\\need_mon.xlsx', sheet_name = 'Sheet 1', index = False)

#******************************************************************************
#need_mon now contains all of the jobs tha need a moniter
#******************************************************************************

#resetting the dataframe
df = initial_df

#Creating a boolean column stating if the job name contains both "aOH" and "mon-box"
df['contains_aOH_mon_box'] = False
for i in range(0,len(df.start_mins)):
    match1 = re.search(r'\baOH\b', df.job_name[i])
    match2 = re.search(r'\bmon-box\b', df.job_name[i])
    if match1 and match2:
        print('match')
        print(df.job_name[i])
        df.contains_aOH_mon_box[i] = True
        
has_mon = df[df.contains_aOH_mon_box == True]

has_mon.to_excel('C:\\Users\\neilklenk\\Documents\\Autosys\\has_mon.xlsx', sheet_name = 'Sheet 1', index = False)
#******************************************************************************
#has_mon continas all of the jobs that already have a moniter on them
#******************************************************************************

#open Excel sheet containing JILs to be made
Job_handle = openpyxl.load_workbook('C:\\Users\\neilklenk\\Documents\\Autosys\\need_mon.xlsx')
Sheet = Job_handle.get_sheet_names()
my_sheet = Job_handle.get_sheet_by_name(Sheet[0]) #need to make sheet name dependent on passed workbook

num_jobs = my_sheet.max_row - 1
num_jobs_iter = my_sheet.max_row + 1

#Collect all of the necessary information into lists 
job_count = 0
for i in range(2,num_jobs_iter):
    #only proceede if an email address exists 
    print(str(my_sheet.cell(row = i, column = app_name_index).value))        
    if str(my_sheet.cell(row = i, column = app_name_index).value) in email_dict: 
       
       #detemine what the name is
       Acr_name_list = (my_sheet.cell(row = i, column = job_name_index).value)
       Acr_name = Acr_name_list.split('-')
       my_Acr_name_build = []
       for j in range (2, len(Acr_name)):    
           my_Acr_name_build.append(Acr_name[j])
           if j < (len(Acr_name)-1):
               my_Acr_name_build.append('-')
           last_word = int(j)
       my_Acr_name = ''.join(my_Acr_name_build)
       del my_Acr_name_build[:] 
        #if name is too long
       len_total = len(my_Acr_name)
       if (len_total > 33):
           my_Acr_name_build = []
           must_remove = ((len_total + 3) - 33)        
           Acr_name_string = str(my_Acr_name)
           Acr_name_index = list(Acr_name_string)
           for j in range(0, (len_total - must_remove)):
               my_Acr_name_build.append(Acr_name_index[j])
           my_Acr_name_build.append(str(i))
           my_Acr_name = ''.join(my_Acr_name_build)
       acronym_job_name.append(my_Acr_name)
       #print(my_Acr_name)
       
       region_ASMS_acronym_job_name.append(my_sheet.cell(row = i, column = job_name_index).value)
       
       #Get the days of week that he job will run
       my_days_of_week = list_builder(i, days_of_week_index)
       days_of_week.append(my_days_of_week)
       #print(my_days_of_week)
       
       #get the start minutes for the job
       my_start_min =  list_builder(i, start_mins_index)
       start_mins.append(my_start_min)
       #print(my_start_min)
       
       #get the job's run window
       my_run_window = list_builder(i, run_window_index)
       run_window.append(my_run_window)
       #print(my_run_window)
       
       #Get the job's run calendar
       my_run_cal = list_builder(i, run_calendar_index)
       run_cal.append(my_run_cal)
       
       #Get the job's exclusion calendar
       my_exc_cal = list_builder(i, exclude_calendar_index)
       exc_cal.append(my_exc_cal)
       
       #determining the sleep frequencey for each aOH job
       my_mins_list = str(my_sheet.cell(row = i, column = start_mins_index).value)
       my_mins = my_mins_list.split(',')
       if my_mins[0] == 'None':
           sleep_freq.append(" ")
       else:
           sleep_time = sleep_freq_calc(my_mins)            
           sleep_freq.append(sleep_time)
       
       #gathering names to match to keys in the email dict
       my_email_list = str(my_sheet.cell(row = i, column = app_name_index).value)
       email_list.append(email_dict[my_email_list])
       
       job_count += 1

#creating the JIL_out file. Writing over portions in the template for each job   
for i in range (0, job_count):
    new_template = Template_contents
    for line in new_template:
        line = write(r'<acronym-job-name>', r'<acronym-job-name>', acronym_job_name[i], line)
        line = write(r'<region-ASMS-acronym-job-name>', r'<region-ASMS-acronym-job-name>', region_ASMS_acronym_job_name[i], line)
        line = write(r'\bdays_of_week\b', r'<copied from job>', days_of_week[i], line)
        line = write(r'\bstart_mins\b', r'<copied from job>', start_mins[i], line)
        line = write(r'\brun_window\b', r'<copied from job>', '\"' + run_window[i] + '\"', line)
        line = write(r'\brun_calendar\b', r'<copied from job>', run_cal[i], line)
        line = write(r'\bexclude_calendar\b', r'<copied from job>', exc_cal[i], line)
        line = write(r'<updated per job frequency>', r'<updated per job frequency>', str(int(sleep_freq[i])), line)
        line = write(r'\bnotification_emailaddress\b', r'<assignment group E1 email@gm.com>', str(email_list[i]), line)
        if line == 'skip me':
            continue
        else:
            #print(line)
            JIL_out.writelines(line)
            
#creating the backout JIL file
for i in range(0,job_count):
    Backout_JIL_out.write("delete_box: gbl-44091-autosys-aOH-" + acronym_job_name[i] + "-mon-box" + '\n')
    
#deleting lists
del acronym_job_name
del region_ASMS_acronym_job_name
del days_of_week 
del start_mins
del run_window
del email_list
del sleep_freq
del run_cal
del exc_cal

JIL_out.close()
Backout_JIL_out.close()
